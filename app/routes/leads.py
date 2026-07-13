# -*- coding: utf-8 -*-
"""线索管理路由"""
import json
import os
import re
from datetime import datetime, date
from io import BytesIO

from flask import (
    Blueprint, render_template, request, redirect, url_for, flash, send_file,
    send_from_directory, abort, current_app, Response, jsonify,
)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.extensions import db
from app.models import Lead, Customer, Attachment
from app.config import Config

leads = Blueprint('leads', __name__, url_prefix='/leads')

# 线索表单下拉选项
ANNOUNCEMENT_TYPE_CHOICES = [
    '公开招标', '中标公告', '成交公告', '更正公告',
    '终止公告', '竞争性磋商', '竞争性谈判', '询价公告',
    '单一来源', '其他公告',
]
SOURCE_TYPE_CHOICES = ['ccgp', 'gdgpo', 'eia', '手动录入']


def _resolve_instance_file(relative_path):
    """将存储在DB中的相对路径解析为 instance 目录下的安全绝对路径

    校验解析结果仍位于 instance 目录内，防止路径穿越。
    返回 (目录, 文件名)，非法或不存在时返回 (None, None)。
    """
    if not relative_path:
        return None, None
    instance_root = os.path.abspath(current_app.instance_path)
    full_path = os.path.abspath(os.path.join(instance_root, relative_path))
    if os.path.commonpath([instance_root, full_path]) != instance_root:
        return None, None
    if not os.path.isfile(full_path):
        return None, None
    return os.path.split(full_path)


_SNAPSHOT_HEAD_INJECT = (
    '<meta charset="utf-8"><meta name="viewport" content="width=device-width, initial-scale=1">'
    '<style>html,body{margin:0;padding:0;background:#f2f1ef;'
    'font-family:-apple-system,BlinkMacSystemFont,"Segoe UI","Microsoft YaHei",sans-serif}</style>'
)


def _wrap_snapshot_html(raw_html, back_url):
    """给爬虫保存的原始快照HTML套一层展示用的外壳（页边距/卡片背景/返回链接），
    仅在渲染时包装，不修改磁盘上存储的原始文件。快照HTML本身常常不是规范的完整
    文档（缺少<meta charset>甚至闭合标签），依赖浏览器的隐式闭合来正常渲染，
    因此这里只在开头标签处插入内容，不手动补齐结尾标签。
    """
    html = raw_html
    if re.search(r'(?i)<html[^>]*>', html):
        html = re.sub(r'(?i)(<html[^>]*>)', r'\1' + _SNAPSHOT_HEAD_INJECT, html, count=1)
    else:
        html = _SNAPSHOT_HEAD_INJECT + html

    toolbar = (
        '<div style="max-width:900px;margin:0 auto;padding:20px 16px 0;font-size:13px">'
        '<a href="%s" style="color:#78716c;text-decoration:none">&larr; 返回线索详情</a></div>'
        '<div style="max-width:900px;margin:12px auto 60px;background:#fff;padding:32px 40px;'
        'border-radius:8px;box-shadow:0 1px 3px rgba(0,0,0,.12)">'
    ) % back_url
    if re.search(r'(?i)<body[^>]*>', html):
        html = re.sub(r'(?i)(<body[^>]*>)', r'\1' + toolbar, html, count=1)
    else:
        html = toolbar + html
    return html


def _parse_date(value):
    """将表单日期字符串转为 date 对象，无效或为空时返回 None"""
    value = (value or '').strip()
    if not value:
        return None
    try:
        return datetime.strptime(value, '%Y-%m-%d').date()
    except ValueError:
        return None


def _parse_budget(value):
    """将表单金额字符串转为 float，无效或为空时返回 None"""
    value = (value or '').strip()
    if not value:
        return None
    try:
        return float(value)
    except (ValueError, TypeError):
        return None


def _apply_filters(query, is_favorited=None):
    """公共方法：对查询对象应用筛选条件"""
    # 默认排除已软删除的线索
    query = query.filter(Lead.deleted == False)  # noqa: E712

    source_type = request.args.get('source_type', '', type=str)
    is_converted = request.args.get('is_converted', '', type=str)
    q = request.args.get('q', '', type=str).strip()
    date_from = request.args.get('date_from', '', type=str)
    date_to = request.args.get('date_to', '', type=str)
    announcement_type = request.args.get('announcement_type', '', type=str)
    region = request.args.get('region', '', type=str)

    if source_type:
        query = query.filter(Lead.source_type == source_type)

    if announcement_type:
        query = query.filter(Lead.announcement_type == announcement_type)

    if region:
        query = query.filter(Lead.region == region)

    if is_favorited == '1':
        query = query.filter(Lead.is_favorited == True)  # noqa: E712

    if is_converted == '0':
        query = query.filter(Lead.is_converted == False)  # noqa: E712
    elif is_converted == '1':
        query = query.filter(Lead.is_converted == True)  # noqa: E712

    if q:
        keyword = f'%{q}%'
        query = query.filter(
            db.or_(
                Lead.project_name.like(keyword),
                Lead.buyer_name.like(keyword),
                Lead.contact_person.like(keyword),
            )
        )

    if date_from:
        try:
            d_from = datetime.strptime(date_from, '%Y-%m-%d').date()
            query = query.filter(Lead.publish_date >= d_from)
        except ValueError:
            pass

    if date_to:
        try:
            d_to = datetime.strptime(date_to, '%Y-%m-%d').date()
            query = query.filter(Lead.publish_date <= d_to)
        except ValueError:
            pass

    return query


@leads.route('/')
def index():
    """线索列表 - 支持筛选、搜索、分页、排序"""
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', Config.PER_PAGE, type=int)
    source_type = request.args.get('source_type', '', type=str)
    is_converted = request.args.get('is_converted', '', type=str)
    q = request.args.get('q', '', type=str).strip()
    date_from = request.args.get('date_from', '', type=str)
    date_to = request.args.get('date_to', '', type=str)
    announcement_type = request.args.get('announcement_type', '', type=str)
    region = request.args.get('region', '', type=str)
    is_favorited = request.args.get('is_favorited')
    sort = request.args.get('sort', 'publish_date', type=str)
    order = request.args.get('order', 'desc', type=str)

    # 白名单排序字段，防止 SQL 注入
    SORT_WHITELIST = {
        'id': Lead.id,
        'publish_date': Lead.publish_date,
        'deadline': Lead.deadline,
        'budget_amount': Lead.budget_amount,
        'created_at': Lead.created_at,
    }
    sort_col = SORT_WHITELIST.get(sort, Lead.publish_date)
    is_desc = order == 'desc'

    query = _apply_filters(Lead.query, is_favorited=is_favorited)

    # 日期/可空字段 null 值排在最后（升序）或最前（降序）
    if sort in ('publish_date', 'deadline'):
        if is_desc:
            query = query.order_by(sort_col.desc().nullslast())
        else:
            query = query.order_by(sort_col.asc().nullslast())
    elif sort == 'budget_amount':
        if is_desc:
            query = query.order_by(sort_col.desc().nullslast())
        else:
            query = query.order_by(sort_col.asc().nullslast())
    else:
        if is_desc:
            query = query.order_by(sort_col.desc())
        else:
            query = query.order_by(sort_col.asc())

    pagination = query.paginate(page=page, per_page=per_page, error_out=False)

    # 供筛选下拉框使用的公告类型/地域候选值（去重后按现有数据动态生成）
    announcement_types = [r[0] for r in db.session.query(Lead.announcement_type)
                          .filter(Lead.announcement_type.isnot(None), Lead.announcement_type != '')
                          .distinct().order_by(Lead.announcement_type).all()]
    regions = [r[0] for r in db.session.query(Lead.region)
              .filter(Lead.region.isnot(None), Lead.region != '')
              .distinct().order_by(Lead.region).all()]

    # 构建查询参数用于分页链接
    query_args = {}
    if source_type:
        query_args['source_type'] = source_type
    if is_converted:
        query_args['is_converted'] = is_converted
    if q:
        query_args['q'] = q
    if date_from:
        query_args['date_from'] = date_from
    if date_to:
        query_args['date_to'] = date_to
    if announcement_type:
        query_args['announcement_type'] = announcement_type
    if region:
        query_args['region'] = region
    if is_favorited:
        query_args['is_favorited'] = is_favorited
    query_args['per_page'] = per_page

    return render_template('leads/list.html',
                           leads=pagination,
                           source_type=source_type,
                           is_converted=is_converted,
                           q=q,
                           date_from=date_from,
                           date_to=date_to,
                           announcement_type=announcement_type,
                           region=region,
                           is_favorited=is_favorited,
                           announcement_types=announcement_types,
                           regions=regions,
                           query_args=query_args,
                           today=date.today(),
                           sort=sort,
                           order=order)


@leads.route('/<int:id>')
def detail(id):
    """线索详情页"""
    lead = Lead.query.get_or_404(id)

    # 格式化原始数据JSON
    raw_data_json = ''
    if lead.raw_data:
        try:
            parsed = json.loads(lead.raw_data)
            raw_data_json = json.dumps(parsed, ensure_ascii=False, indent=2)
        except (json.JSONDecodeError, TypeError):
            raw_data_json = lead.raw_data

    return render_template('leads/detail.html',
                           lead=lead,
                           raw_data_json=raw_data_json,
                           today=date.today())


@leads.route('/<int:id>/snapshot')
def snapshot(id):
    """查看详情页HTML快照（本地留档，公告在官网被撤回/修改后仍可查看原文）"""
    lead = Lead.query.get_or_404(id)
    directory, filename = _resolve_instance_file(lead.html_snapshot_path)
    if not directory:
        abort(404)
    with open(os.path.join(directory, filename), encoding='utf-8') as f:
        raw_html = f.read()
    wrapped = _wrap_snapshot_html(raw_html, url_for('leads.detail', id=lead.id))
    return Response(wrapped, mimetype='text/html')


@leads.route('/<int:id>/attachments/<int:attachment_id>')
def download_attachment(id, attachment_id):
    """下载线索详情页附件"""
    attachment = Attachment.query.filter_by(id=attachment_id, lead_id=id).first_or_404()
    directory, filename = _resolve_instance_file(attachment.local_path)
    if not directory:
        abort(404)
    return send_from_directory(directory, filename, as_attachment=True,
                                download_name=attachment.file_name or filename)


@leads.route('/<int:id>/convert', methods=['POST'])
def convert(id):
    """一键转客户 - 将线索转为Customer记录"""
    lead = Lead.query.get_or_404(id)

    if lead.is_converted:
        flash('该线索已转化，无需重复操作', 'warning')
        return redirect(url_for('leads.detail', id=id))

    try:
        customer = Customer(
            company_name=lead.buyer_name or '未知单位',
            contact_person=lead.contact_person or '',
            phone=lead.phone or '',
            source='爬虫采集',
            status='新线索',
            notes=(f'由线索#{lead.id}自动转化\n'
                   f'项目名称：{lead.project_name or ""}\n'
                   f'招标编号：{lead.bidding_number or ""}'),
        )
        db.session.add(customer)
        db.session.flush()  # 获取 customer.id

        lead.is_converted = True
        lead.converted_customer_id = customer.id

        db.session.commit()
        flash(f'成功转化为客户：{customer.company_name}', 'success')
        return redirect(url_for('customers.detail', id=customer.id))
    except Exception:
        db.session.rollback()
        flash('转化失败，请重试或联系管理员', 'danger')
        return redirect(url_for('leads.detail', id=id))


@leads.route('/batch_convert', methods=['POST'])
def batch_convert():
    """批量转客户"""
    lead_ids = request.form.getlist('lead_ids')

    if not lead_ids:
        flash('请先选择要转化的线索', 'warning')
        return redirect(url_for('leads.index'))

    success_count = 0
    skip_count = 0

    for lid in lead_ids:
        try:
            lead = Lead.query.get(int(lid))
            if not lead or lead.is_converted:
                skip_count += 1
                continue

            customer = Customer(
                company_name=lead.buyer_name or '未知单位',
                contact_person=lead.contact_person or '',
                phone=lead.phone or '',
                source='爬虫采集',
                status='新线索',
                notes=(f'由线索#{lead.id}批量转化\n'
                       f'项目名称：{lead.project_name or ""}\n'
                       f'招标编号：{lead.bidding_number or ""}'),
            )
            db.session.add(customer)
            db.session.flush()

            lead.is_converted = True
            lead.converted_customer_id = customer.id
            success_count += 1
        except Exception:
            db.session.rollback()
            skip_count += 1
            continue

    try:
        db.session.commit()
        if success_count:
            flash(f'成功转化 {success_count} 条线索为客户', 'success')
        if skip_count:
            flash(f'{skip_count} 条线索已跳过（已转化或不存在）', 'info')
    except Exception:
        db.session.rollback()
        flash('批量转化失败，请重试或联系管理员', 'danger')

    return redirect(url_for('leads.index'))


@leads.route('/export')
def export():
    """导出Excel - 导出当前筛选条件下的所有线索"""
    query = _apply_filters(Lead.query, is_favorited=request.args.get('is_favorited'))
    leads_list = query.order_by(Lead.publish_date.desc().nullslast()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = '线索列表'

    # 样式定义
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    # 写表头
    headers = ['序号', '招标编号', '项目名称', '公告类型', '采购单位', '地域', '联系人', '电话',
               '预算金额(元)', '发布日期', '截止日期', '来源', '状态', '来源URL']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # 写数据
    for row_idx, lead in enumerate(leads_list, 2):
        row_data = [
            row_idx - 1,
            lead.bidding_number or '',
            lead.project_name or '',
            lead.announcement_type or '',
            lead.buyer_name or '',
            lead.region or '',
            lead.contact_person or '',
            lead.phone or '',
            lead.budget_amount if lead.budget_amount is not None else '',
            lead.publish_date.strftime('%Y-%m-%d') if lead.publish_date else '',
            lead.deadline.strftime('%Y-%m-%d') if lead.deadline else '',
            lead.source_type or '',
            '已转化' if lead.is_converted else '未转化',
            lead.source_url or '',
        ]
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center', wrap_text=True)

    # 自动列宽
    col_widths = [6, 22, 42, 12, 25, 8, 12, 15, 15, 13, 13, 10, 10, 42]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # 冻结首行
    ws.freeze_panes = 'A2'

    # 输出到内存
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f'线索列表_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@leads.route('/new', methods=['GET', 'POST'])
def create():
    """手动新增线索"""
    if request.method == 'POST':
        project_name = request.form.get('project_name', '').strip()
        if not project_name:
            flash('项目名称不能为空', 'error')
            return redirect(url_for('leads.create'))

        lead = Lead(
            project_name=project_name,
            bidding_number=request.form.get('bidding_number', '').strip() or None,
            announcement_type=request.form.get('announcement_type', ''),
            buyer_name=request.form.get('buyer_name', '').strip(),
            buyer_address=request.form.get('buyer_address', '').strip(),
            region=request.form.get('region', '').strip(),
            contact_person=request.form.get('contact_person', '').strip(),
            phone=request.form.get('phone', '').strip(),
            agency_name=request.form.get('agency_name', '').strip(),
            agency_phone=request.form.get('agency_phone', '').strip(),
            budget_amount=_parse_budget(request.form.get('budget_amount', '')),
            publish_date=_parse_date(request.form.get('publish_date', '')),
            publish_time=request.form.get('publish_time', '').strip(),
            deadline=_parse_date(request.form.get('deadline', '')),
            source_url=request.form.get('source_url', '').strip(),
            source_type=request.form.get('source_type', '手动录入'),
        )
        db.session.add(lead)
        try:
            db.session.commit()
            flash('线索添加成功', 'success')
        except Exception:
            db.session.rollback()
            flash('操作失败，请重试（招标编号可能重复）', 'error')
            return redirect(url_for('leads.create'))
        return redirect(url_for('leads.detail', id=lead.id))

    return render_template(
        'leads/form.html',
        lead=None,
        announcement_type_choices=ANNOUNCEMENT_TYPE_CHOICES,
        source_type_choices=SOURCE_TYPE_CHOICES,
    )


@leads.route('/<int:id>/edit', methods=['GET', 'POST'])
def edit(id):
    """编辑线索"""
    lead = Lead.query.get_or_404(id)

    if request.method == 'POST':
        project_name = request.form.get('project_name', '').strip()
        if not project_name:
            flash('项目名称不能为空', 'error')
            return redirect(url_for('leads.edit', id=id))

        lead.project_name = project_name
        bidding_number = request.form.get('bidding_number', '').strip()
        lead.bidding_number = bidding_number if bidding_number else None
        lead.announcement_type = request.form.get('announcement_type', '')
        lead.buyer_name = request.form.get('buyer_name', '').strip()
        lead.buyer_address = request.form.get('buyer_address', '').strip()
        lead.region = request.form.get('region', '').strip()
        lead.contact_person = request.form.get('contact_person', '').strip()
        lead.phone = request.form.get('phone', '').strip()
        lead.agency_name = request.form.get('agency_name', '').strip()
        lead.agency_phone = request.form.get('agency_phone', '').strip()
        lead.budget_amount = _parse_budget(request.form.get('budget_amount', ''))
        lead.publish_date = _parse_date(request.form.get('publish_date', ''))
        lead.publish_time = request.form.get('publish_time', '').strip()
        lead.deadline = _parse_date(request.form.get('deadline', ''))
        lead.source_url = request.form.get('source_url', '').strip()
        lead.source_type = request.form.get('source_type', '手动录入')
        try:
            db.session.commit()
            flash('线索信息已更新', 'success')
        except Exception:
            db.session.rollback()
            flash('操作失败，请重试（招标编号可能重复）', 'error')
            return redirect(url_for('leads.edit', id=id))
        return redirect(url_for('leads.detail', id=lead.id))

    return render_template(
        'leads/form.html',
        lead=lead,
        announcement_type_choices=ANNOUNCEMENT_TYPE_CHOICES,
        source_type_choices=SOURCE_TYPE_CHOICES,
    )


@leads.route('/<int:id>/delete', methods=['POST'])
def delete(id):
    """软删除线索（将deleted标记置为True）"""
    lead = Lead.query.get_or_404(id)
    lead.deleted = True
    try:
        db.session.commit()
        flash('线索已删除', 'success')
    except Exception:
        db.session.rollback()
        flash('操作失败，请重试', 'error')
    return redirect(url_for('leads.index'))


@leads.route('/<int:lead_id>/favorite', methods=['POST'])
def favorite(lead_id):
    """切换线索收藏状态"""
    lead = Lead.query.get_or_404(lead_id)
    lead.is_favorited = not lead.is_favorited
    try:
        db.session.commit()
    except Exception:
        db.session.rollback()
        # AJAX 请求返回 JSON 错误
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'success': False, 'message': '操作失败'}), 500
        flash('操作失败', 'error')
        return redirect(url_for('leads.detail', id=lead_id))

    # AJAX 请求返回 JSON
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return jsonify({'success': True, 'is_favorited': lead.is_favorited})

    # 传统表单提交，重定向回详情页
    flash('已取消收藏' if not lead.is_favorited else '已收藏', 'success')
    return redirect(url_for('leads.detail', id=lead_id))
