from datetime import date, datetime
from io import BytesIO

from flask import (
    Blueprint, render_template, redirect, url_for,
    request, flash, jsonify, send_file,
)
from sqlalchemy import or_
from openpyxl import Workbook, load_workbook

from app.models import Customer, FollowUp, Lead
from app.extensions import db
from app.config import Config

customers = Blueprint('customers', __name__, url_prefix='/customers')

# ---- 常量：下拉选项 ----
STATUS_CHOICES = ['新线索', '跟进中', '已报价', '已成交', '已流失']
INDUSTRY_CHOICES = ['市政工程', '建筑工程', '装饰装修', '水务', '电力', '通信', '经销商/建材', '其他']
SOURCE_CHOICES = ['爬虫采集', '地图POI采集', '手动录入', 'Excel导入', '转介绍', '其他']
FOLLOW_TYPE_CHOICES = ['电话', '拜访', '微信', '邮件']

# Excel 导入时表头与字段名的映射
EXCEL_HEADER_MAP = {
    '公司名称': 'company_name',
    '联系人': 'contact_person',
    '电话': 'phone',
    '地址': 'address',
    '行业类型': 'industry_type',
    '行业': 'industry_type',
    '来源': 'source',
    '状态': 'status',
    '备注': 'notes',
}


@customers.route('/')
def index():
    """客户列表：支持状态/行业/来源筛选 + 搜索 + 分页"""
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', Config.PER_PAGE, type=int)
    status = request.args.get('status', '')
    industry = request.args.get('industry', '')
    source = request.args.get('source', '')
    q = request.args.get('q', '').strip()

    query = Customer.query
    if status:
        query = query.filter(Customer.status == status)
    if industry:
        query = query.filter(Customer.industry_type == industry)
    if source:
        query = query.filter(Customer.source == source)
    if q:
        query = query.filter(or_(
            Customer.company_name.contains(q),
            Customer.contact_person.contains(q),
            Customer.phone.contains(q),
        ))

    pagination = query.order_by(Customer.created_at.desc()).paginate(
        page=page, per_page=per_page, error_out=False,
    )

    return render_template(
        'customers/list.html',
        customers=pagination.items,
        pagination=pagination,
        status=status, industry=industry, source=source, q=q,
        status_choices=STATUS_CHOICES,
        industry_choices=INDUSTRY_CHOICES,
        source_choices=SOURCE_CHOICES,
    )


@customers.route('/new', methods=['GET', 'POST'])
def create():
    """新增客户"""
    if request.method == 'POST':
        company_name = request.form.get('company_name', '').strip()
        if not company_name:
            flash('公司名称不能为空', 'error')
            return redirect(url_for('customers.create'))

        customer = Customer(
            company_name=company_name,
            contact_person=request.form.get('contact_person', '').strip(),
            phone=request.form.get('phone', '').strip(),
            address=request.form.get('address', '').strip(),
            industry_type=request.form.get('industry_type', ''),
            source=request.form.get('source', ''),
            status=request.form.get('status', '新线索'),
            notes=request.form.get('notes', '').strip(),
        )
        db.session.add(customer)
        try:
            db.session.commit()
            flash('客户添加成功', 'success')
        except Exception:
            db.session.rollback()
            flash('操作失败，请重试', 'error')
            return redirect(url_for('customers.create'))
        return redirect(url_for('customers.detail', id=customer.id))

    return render_template(
        'customers/form.html',
        customer=None,
        status_choices=STATUS_CHOICES,
        industry_choices=INDUSTRY_CHOICES,
        source_choices=SOURCE_CHOICES,
    )


@customers.route('/<int:id>')
def detail(id):
    """客户详情：基本信息 + 跟进记录"""
    customer = Customer.query.get_or_404(id)
    followups = customer.follow_ups.order_by(FollowUp.created_at.desc()).all()
    return render_template(
        'customers/detail.html',
        customer=customer,
        followups=followups,
        today=date.today(),
        status_choices=STATUS_CHOICES,
        follow_type_choices=FOLLOW_TYPE_CHOICES,
    )


@customers.route('/<int:id>/edit', methods=['GET', 'POST'])
def edit(id):
    """编辑客户"""
    customer = Customer.query.get_or_404(id)

    if request.method == 'POST':
        company_name = request.form.get('company_name', '').strip()
        if not company_name:
            flash('公司名称不能为空', 'error')
            return redirect(url_for('customers.edit', id=id))

        customer.company_name = company_name
        customer.contact_person = request.form.get('contact_person', '').strip()
        customer.phone = request.form.get('phone', '').strip()
        customer.address = request.form.get('address', '').strip()
        customer.industry_type = request.form.get('industry_type', '')
        customer.source = request.form.get('source', '')
        customer.status = request.form.get('status', '新线索')
        customer.notes = request.form.get('notes', '').strip()
        try:
            db.session.commit()
            flash('客户信息已更新', 'success')
        except Exception:
            db.session.rollback()
            flash('操作失败，请重试', 'error')
            return redirect(url_for('customers.edit', id=id))
        return redirect(url_for('customers.detail', id=customer.id))

    return render_template(
        'customers/form.html',
        customer=customer,
        status_choices=STATUS_CHOICES,
        industry_choices=INDUSTRY_CHOICES,
        source_choices=SOURCE_CHOICES,
    )


@customers.route('/<int:id>/delete', methods=['POST'])
def delete(id):
    """删除客户（级联删除跟进记录）"""
    customer = Customer.query.get_or_404(id)
    # 清理引用该客户的 Lead 记录
    Lead.query.filter_by(converted_customer_id=id).update({
        'is_converted': False,
        'converted_customer_id': None,
    })
    db.session.delete(customer)
    try:
        db.session.commit()
        flash('客户已删除', 'success')
    except Exception:
        db.session.rollback()
        flash('操作失败，请重试', 'error')
    return redirect(url_for('customers.index'))


@customers.route('/<int:id>/followup', methods=['POST'])
def followup(id):
    """添加跟进记录"""
    customer = Customer.query.get_or_404(id)
    content = request.form.get('content', '').strip()
    if not content:
        flash('跟进内容不能为空', 'error')
        return redirect(url_for('customers.detail', id=id))

    next_date_str = request.form.get('next_date', '').strip()
    next_date = None
    if next_date_str:
        try:
            next_date = datetime.strptime(next_date_str, '%Y-%m-%d').date()
        except ValueError:
            flash('下次跟进日期格式不正确', 'error')
            return redirect(url_for('customers.detail', id=id))

    fu = FollowUp(
        customer_id=id,
        content=content,
        follow_type=request.form.get('follow_type', '电话'),
        next_action=request.form.get('next_action', '').strip(),
        next_date=next_date,
    )
    db.session.add(fu)
    try:
        db.session.commit()
        flash('跟进记录已添加', 'success')
    except Exception:
        db.session.rollback()
        flash('操作失败，请重试', 'error')
    return redirect(url_for('customers.detail', id=id))


@customers.route('/<int:id>/status', methods=['POST'])
def update_status(id):
    """AJAX 快速修改客户状态"""
    customer = Customer.query.get_or_404(id)
    new_status = request.form.get('status', '').strip()
    if new_status not in STATUS_CHOICES:
        return jsonify({'success': False, 'message': '无效的状态'}), 400

    customer.status = new_status
    try:
        db.session.commit()
    except Exception:
        db.session.rollback()
        return jsonify({'success': False, 'message': '操作失败，请重试'}), 500
    return jsonify({'success': True, 'message': '状态已更新', 'status': new_status})


@customers.route('/export')
def export():
    """导出全部客户为 Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = '客户列表'
    headers = ['公司名称', '联系人', '电话', '地址', '行业类型', '来源', '状态', '备注', '创建时间']
    ws.append(headers)

    all_customers = Customer.query.order_by(Customer.created_at.desc()).all()
    for c in all_customers:
        ws.append([
            c.company_name or '',
            c.contact_person or '',
            c.phone or '',
            c.address or '',
            c.industry_type or '',
            c.source or '',
            c.status or '',
            c.notes or '',
            c.created_at.strftime('%Y-%m-%d %H:%M') if c.created_at else '',
        ])

    # 自适应列宽
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f'客户列表_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


@customers.route('/import', methods=['POST'])
def import_excel():
    """导入 Excel 批量创建客户"""
    file = request.files.get('file')
    if not file or not file.filename:
        flash('请选择要导入的Excel文件', 'error')
        return redirect(url_for('customers.index'))

    if not file.filename.endswith('.xlsx'):
        flash('只支持 .xlsx 格式的文件', 'error')
        return redirect(url_for('customers.index'))

    try:
        wb = load_workbook(BytesIO(file.read()))
        ws = wb.active

        # 读取第一行表头，建立列索引映射
        headers = [str(cell.value).strip() if cell.value else '' for cell in ws[1]]
        col_indices = {}
        for i, h in enumerate(headers):
            field = EXCEL_HEADER_MAP.get(h)
            if field and field not in col_indices:
                col_indices[field] = i

        if 'company_name' not in col_indices:
            flash('Excel文件中必须包含"公司名称"列', 'error')
            return redirect(url_for('customers.index'))

        def get_val(row, field):
            idx = col_indices.get(field)
            if idx is not None and idx < len(row):
                val = row[idx]
                return str(val).strip() if val is not None else ''
            return ''

        count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            company_name = get_val(row, 'company_name')
            if not company_name:
                continue

            customer = Customer(
                company_name=company_name,
                contact_person=get_val(row, 'contact_person'),
                phone=get_val(row, 'phone'),
                address=get_val(row, 'address'),
                industry_type=get_val(row, 'industry_type') or '其他',
                source=get_val(row, 'source') or 'Excel导入',
                status=get_val(row, 'status') or '新线索',
                notes=get_val(row, 'notes'),
            )
            db.session.add(customer)
            count += 1

        db.session.commit()
        flash(f'成功导入 {count} 条客户记录', 'success')

    except Exception:
        db.session.rollback()
        flash('导入失败，请重试或联系管理员', 'error')

    return redirect(url_for('customers.index'))
